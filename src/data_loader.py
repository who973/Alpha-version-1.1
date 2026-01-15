import os
import csv
from openpyxl import load_workbook


def load_csv(path: str) -> list:
    data = []
    label_mapping = {"negative": 0, "positive": 1, "neutral": 2}

    with open(path, 'r', encoding='utf-8-sig', newline='') as csvfile:
        reader = csv.DictReader(csvfile)
        reader.fieldnames = [name.strip() for name in reader.fieldnames]

        required_columns = ['text', 'label']
        for col in required_columns:
            if col not in reader.fieldnames:
                raise ValueError(f"Отсутствует обязательный столбец: {col}")

        for idx, row in enumerate(reader):
            text = row['text'].strip()
            if not text or len(text) < 5:
                continue

            label_str = row['label'].strip().lower()
            if label_str not in label_mapping:
                continue  # пропустить неизвестные метки

            data.append({
                'id': idx,
                'text': text,
                'date': row.get('date', "2025-12-26"),
                'label': label_mapping[label_str]
            })

    return data



def load_excel(path):
    data = []
    wb = load_workbook(path, read_only=True)
    ws = wb.active

    headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    if 'text' not in headers or 'label' not in headers:
        raise ValueError("Отсутствует обязательный столбец: text или label")

    text_idx = headers.index('text')
    label_idx = headers.index('label')
    date_idx = headers.index('date') if 'date' in headers else None

    for row in ws.iter_rows(min_row=2):
        data.append({
            'text': str(row[text_idx].value).strip(),
            'label': int(row[label_idx].value),
            'date': str(row[date_idx].value).strip() if date_idx is not None else None
        })
    return data


def load_data(path: str) -> list:
    if not os.path.exists(path):
        raise FileNotFoundError(f"Файл {path} не найден")
    if path.lower().endswith('.csv'):
        raw_data = load_csv(path)
    elif path.lower().endswith(('.xls', '.xlsx')):
        raw_data = load_excel(path)
    else:
        raise ValueError("Неподдерживаемый формат файла")

    validated_data = []
    dropped = 0

    for idx, row in enumerate(raw_data):
        text = row.get('text')
        label = row.get('label')

        if not text or len(text) < 5:
            dropped += 1
            continue

        if label not in (0, 1):
            dropped += 1
            continue

        validated_data.append({
            'id': idx,
            'text': text,
            'label': label,
            'date': row.get('date')
        })

    print(
        f"Загружено: {len(raw_data)}, "
        f"принято: {len(validated_data)}, "
        f"отброшено: {dropped}"
    )

    return validated_data
